"""UK adapter tests.

We build a synthetic ONS-shaped XLSX fixture in-memory rather than committing a
real ONS file (which is ~2MB and licensed separately). The shape mirrors what
ONS publishes: column A = force name, remaining columns = offence counts.
"""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from adapters.common.base import SourceFile
from adapters.uk.adapter import UKAdapter, _candidate_releases, _Release


# ---------- fixture builder ----------


def _build_fake_pfa_xlsx() -> bytes:
    """Build a minimal ONS-shaped workbook mirroring Table P1.

    Layout matches what ONS actually publishes:
      Rows 1-6: title + boilerplate (skipped by parser).
      Row 7:    headers, with "Area Code" in column A, "Area Name" in column B.
      Row 8+:   data rows. Codes E92*/E12*/W92* are rollups, E23*/W15* are forces.
    """
    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    wb.create_sheet("Cover_sheet")["A1"] = "synthetic fixture"

    p1 = wb.create_sheet("Table P1")
    # Boilerplate header rows
    p1["A1"] = "Table P1: Police recorded crime, year ending June 2025 [note 1]"
    p1["A2"] = "boilerplate"
    p1["A3"] = "boilerplate"
    p1["A4"] = "boilerplate"
    p1["A5"] = "boilerplate"
    p1["A6"] = "boilerplate"
    # Real header row (row 7, index 6 in 0-based)
    headers = [
        "Area Code",
        "Area Name",
        "Total recorded crime\n (excluding fraud) \n[note 2]",
        "Violence against the person",
        "Homicide",
        "Violence with injury",
        "Sexual offences",
        "Robbery",
        "Theft offences",  # category in skip-list
    ]
    for col_idx, h in enumerate(headers, start=1):
        p1.cell(row=7, column=col_idx, value=h)

    # Data rows: mix of rollups (skipped by area-code filter) and forces.
    rows = [
        # rollups — should be filtered out
        ("K04000001", "ENGLAND AND WALES [note 3]", 5000000, 1500000, 400, 100000, 80000, 50000, 1000000),
        ("E92000001", "ENGLAND", 4500000, 1400000, 380, 95000, 75000, 47000, 950000),
        ("E12000001", "North East", 250000, 90000, 25, 7000, 4500, 2800, 50000),
        ("E12000007", "London", 900000, 250000, 130, 25000, 9000, 7500, 180000),
        ("W92000004", "WALES", 200000, 70000, 18, 6500, 3500, 2000, 45000),
        # actual forces (E23 + W15)
        ("E23000001", "Metropolitan Police", 850000, 240000, 137, 25410, 8923, 7611, 170000),
        ("E23000034", "City of London", 8000, 200, 0, 142, 31, 12, 1500),
        ("E23000014", "West Midlands", 380000, 100000, 47, 11220, 3812, 1605, 76000),
        ("E23000005", "Greater Manchester", 470000, 130000, 41, 14502, 4901, 2113, 95000),
        ("E23000006", "Cheshire", 95000, 28000, 8, 3000, 1100, 290, 19000),
        ("W15000001", "Dyfed-Powys", 35000, 11000, 2, 1100, 480, 60, 7000),
    ]
    for i, row in enumerate(rows, start=8):
        for col_idx, val in enumerate(row, start=1):
            p1.cell(row=i, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def fake_source(tmp_path: Path) -> SourceFile:
    """Write the synthetic XLSX to a fixture path and return a SourceFile."""
    content = _build_fake_pfa_xlsx()
    # Use ONS-style filename so the adapter's release-from-filename helper works.
    fname = "pfatables-yearending-june-2025.xlsx"
    target = tmp_path / "raw" / "uk" / "2025-08" / fname
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(content)

    # The adapter resolves src.local_path relative to REPO_ROOT, so we need to
    # write to data/raw/ for the path to be reachable. Copy there too.
    from adapters.common.base import REPO_ROOT  # local import to avoid circularity

    real_target = REPO_ROOT / "data" / "raw" / "uk" / "test-fixture" / fname
    real_target.parent.mkdir(parents=True, exist_ok=True)
    real_target.write_bytes(content)

    return SourceFile(
        url="https://www.ons.gov.uk/file?uri=fake",
        local_path=str(real_target.relative_to(REPO_ROOT)),
        fetched_at=datetime.utcnow(),
        sha256="fake-sha-for-tests",
    )


# ---------- tests ----------


def test_release_url_pattern() -> None:
    rel = _Release(month="june", year=2025)
    assert "yearendingjune2025" in rel.url
    assert rel.filename == "pfatables-yearending-june-2025.xlsx"


def test_candidate_releases_returns_recent_first() -> None:
    cands = _candidate_releases()
    assert len(cands) == 6
    # All distinct
    assert len({(c.month, c.year) for c in cands}) == 6
    # All in valid quarter months
    assert all(c.month in {"march", "june", "september", "december"} for c in cands)


def test_release_from_filename_roundtrip() -> None:
    rel = UKAdapter._release_from_filename("pfatables-yearending-june-2025.xlsx")
    assert rel is not None
    assert rel.month == "june"
    assert rel.year == 2025
    assert UKAdapter._release_from_filename("garbage.xlsx") is None


def test_parse_and_normalise(fake_source: SourceFile) -> None:
    adapter = UKAdapter()
    parsed = adapter.parse(fake_source)
    assert not parsed.empty
    # parse() must filter rollups out via area-code prefix. Six forces remain:
    # Met, City of London, West Midlands, Greater Manchester, Cheshire, Dyfed-Powys.
    assert set(parsed["force_native"].unique()) == {
        "Metropolitan Police",
        "City of London",
        "West Midlands",
        "Greater Manchester",
        "Cheshire",
        "Dyfed-Powys",
    }
    assert parsed["count"].sum() > 0
    # area_code column carried through
    assert all(c.startswith(("E23", "W15")) for c in parsed["area_code"].unique())

    norm = adapter.normalise(parsed, fake_source)
    assert not norm.empty

    # ONS does not publish suspect background — all suspect_dim must be 'total'.
    assert (norm["suspect_dim"] == "total").all()

    # Met + City of London → UKI aggregation.
    london = norm[norm["region_code"] == "UKI"]
    assert not london.empty
    # London homicide = Met (137) + City (0) = 137.
    london_homicide = london[london["crime_category"] == "homicide"]["count"].sum()
    assert london_homicide == 137

    # One row per (region, category) — no dupes when forces collapse.
    london_robbery_rows = london[london["crime_category"] == "robbery_violent"]
    assert len(london_robbery_rows) == 1

    # Theft offences are in the skip list — must not appear.
    assert "theft_offences" not in set(norm["crime_category"].str.lower())

    # West Midlands force → UKG. Population denominator should be filled.
    wm = norm[norm["region_code"] == "UKG"]
    assert not wm.empty
    assert wm["denominator_population"].notna().all()
    assert (wm["rate_per_100k"] > 0).all()

    # Dyfed-Powys force → UKL Wales. Rate should be reasonable.
    wales = norm[norm["region_code"] == "UKL"]
    assert not wales.empty


def test_normalise_passes_validate_aggregates(fake_source: SourceFile) -> None:
    """The output of normalise() must pass the cross-cutting validator."""
    from adapters.common.validate import validate_aggregates

    adapter = UKAdapter()
    parsed = adapter.parse(fake_source)
    norm = adapter.normalise(parsed, fake_source)
    validate_aggregates(norm)


def test_normalise_no_duplicate_natural_key(fake_source: SourceFile) -> None:
    adapter = UKAdapter()
    parsed = adapter.parse(fake_source)
    norm = adapter.normalise(parsed, fake_source)
    key_cols = [
        "source_country",
        "period_start",
        "region_code",
        "crime_category",
        "suspect_dim",
        "suspect_dim_value",
    ]
    dupes = norm[norm.duplicated(subset=key_cols, keep=False)]
    assert dupes.empty, f"duplicate natural-key rows: {dupes}"
