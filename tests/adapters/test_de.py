"""DE adapter tests — synthetic BKA Zeitreihen XLSX fixtures.

The DE adapter reads three Zeitreihen workbooks: total suspects (T20),
German suspects (T40), non-German suspects (T50). Each has the layout:
header at row 5 (Schlüssel / Straftat / Jahr / Tatverdächtige insgesamt /
age bands), data from row 11+.
"""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from adapters.common.base import REPO_ROOT, SourceFile
from adapters.de.adapter import DEAdapter


def _build_fake_zr_xlsx(suspect_dim_label: str) -> bytes:
    """Build a Zeitreihen-shaped workbook for one suspect dimension.

    suspect_dim_label is just a tag we encode into the sheet title so we
    can verify the dim guesser works.
    """
    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    sheet = wb.create_sheet(f"T20_ZR insg")  # exact name not important — adapter picks any sheet with "ZR"

    # Boilerplate rows 1-5 (BKA's header block)
    for r in range(1, 6):
        sheet.cell(row=r, column=1, value="boilerplate")
    # Row 6: column headers (the adapter reads with header=5, so row index 6 here
    # = pandas header=5)
    headers = ["Schlüssel", "Straftat", "Jahr", "Tatverdächtige insgesamt", "Kinder"]
    for c, h in enumerate(headers, start=1):
        sheet.cell(row=6, column=c, value=h)
    # Rows 7-10: sub-headers / spurious rows
    for r in range(7, 11):
        sheet.cell(row=r, column=1, value=None)
    # Row 11+: data
    rows = [
        ("------", "Straftaten insgesamt", 2024, 1_500_000, 1000),
        ("------", "Straftaten insgesamt", 2023, 1_400_000, 900),
        ("010000", "Mord § 211 StGB", 2024, 800, 0),
        ("010000", "Mord § 211 StGB", 2023, 750, 0),
        ("020000", "Totschlag § 212 StGB", 2024, 1500, 5),
        ("020000", "Totschlag § 212 StGB", 2023, 1400, 4),
        ("110000", "Vergewaltigung u. sex. Nötigung", 2024, 15000, 5),
        ("210000", "Raub", 2024, 25000, 50),
        ("222000", "Gefährliche und schwere KV", 2024, 150000, 200),
        ("892000", "Gewaltkriminalität", 2024, 200000, 250),
        # Should be skipped (not in our schluessel map)
        ("500000", "Diebstahl", 2024, 1_000_000, 5000),
    ]
    for ri, row in enumerate(rows, start=11):
        for ci, val in enumerate(row, start=1):
            sheet.cell(row=ri, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def fake_t20_source() -> SourceFile:
    content = _build_fake_zr_xlsx("total")
    target = REPO_ROOT / "data" / "raw" / "de" / "test-fixture" / "T20-tv-insg.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(content)
    return SourceFile(
        url="manual",
        local_path=str(target.relative_to(REPO_ROOT)),
        fetched_at=datetime.utcnow(),
        sha256="fake-t20-sha",
    )


def test_parse_skips_sentinel_rows(fake_t20_source: SourceFile) -> None:
    a = DEAdapter()
    df = a.parse(fake_t20_source)
    # "------" sentinel rows must be dropped (Schlüssel filter is ^\d{6}$).
    assert (df["schluessel"] != "------").all()
    # Diebstahl (500000) is present at parse stage; it's filtered later at normalise.
    assert "500000" in set(df["schluessel"])
    # Suspect dim inferred from filename (T20-tv-insg.xlsx → "total").
    assert (df["suspect_dim"] == "total").all()


def test_normalise_collapses_homicide(fake_t20_source: SourceFile) -> None:
    a = DEAdapter()
    df = a.normalise(a.parse(fake_t20_source), fake_t20_source)
    # Mord (010000) + Totschlag (020000) both map to homicide → collapsed.
    homicide = df[df["crime_category"] == "homicide"]
    assert len(homicide) == 1
    # 800 (Mord 2024) + 1500 (Totschlag 2024) = 2300
    assert int(homicide["count"].iloc[0]) == 2300
    # Diebstahl is filtered out.
    assert "theft" not in set(df["crime_category"].str.lower())


def test_normalise_emits_nuts0(fake_t20_source: SourceFile) -> None:
    a = DEAdapter()
    df = a.normalise(a.parse(fake_t20_source), fake_t20_source)
    assert (df["region_code"] == "DE").all()
    assert (df["region_level"] == 0).all()


def test_filename_dim_inference() -> None:
    assert DEAdapter._guess_dim_from_filename("T20-tv-insg.xlsx") == "total"
    assert DEAdapter._guess_dim_from_filename("T40-tv-deutsch.xlsx") == "national"
    assert DEAdapter._guess_dim_from_filename("T50-tv-nichtdeutsch.xlsx") == "foreign"
